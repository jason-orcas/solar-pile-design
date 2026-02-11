"""TOPL (Top of Pile Load) document parser for solar tracker manufacturers.

Extracts project info and unfactored per-pile loads from ATI, Nevados,
and Nextpower TOPL documents (PDF or XLSX).  Produces a standardised
result that can be mapped directly to SPORK session-state keys.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from enum import Enum
from typing import Any


# ============================================================================
# Enums & Data Model
# ============================================================================

class Manufacturer(Enum):
    ATI = "ATI (Array Technologies)"
    NEVADOS = "Nevados"
    NEXTPOWER = "Nextpower"


@dataclass
class TOPLProjectInfo:
    """Project metadata extracted from a TOPL document."""
    project_name: str = ""
    project_number: str = ""
    client: str = ""
    location: str = ""
    manufacturer: str = ""
    asce_version: str = ""
    wind_speed_mph: float | None = None
    ground_snow_psf: float | None = None
    seismic_sds: float | None = None
    risk_category: str = ""
    exposure: str = ""


@dataclass
class TOPLColumnLoads:
    """Per-pile unfactored loads in SPORK units.

    All forces in lbs, all moments in ft-lbs.  Values are positive
    (absolute magnitudes); sign conventions handled at extraction.
    """
    column_type: str = ""
    dead_load: float = 0.0        # lbs (compression)
    snow_load: float = 0.0        # lbs (compression)
    wind_up: float = 0.0          # lbs (uplift)
    wind_down: float = 0.0        # lbs (downforce)
    wind_lateral: float = 0.0     # lbs
    wind_moment: float = 0.0      # ft-lbs
    seismic_lateral: float = 0.0  # lbs
    seismic_vertical: float = 0.0 # lbs
    seismic_moment: float = 0.0   # ft-lbs
    lever_arm: float = 4.0        # ft (column height above grade)


@dataclass
class TOPLParseResult:
    """Complete result of parsing a TOPL document."""
    success: bool
    manufacturer: Manufacturer
    project_info: TOPLProjectInfo
    column_options: list[str] = field(default_factory=list)
    loads_by_column: dict[str, TOPLColumnLoads] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


# ============================================================================
# Public API
# ============================================================================

def parse_topl(
    file_bytes: bytes,
    filename: str,
    manufacturer: Manufacturer,
) -> TOPLParseResult:
    """Parse a TOPL document and return extracted data.

    Args:
        file_bytes: Raw file contents.
        filename: Original filename (used for extension check).
        manufacturer: Which manufacturer's format to expect.

    Returns:
        TOPLParseResult with extracted project info and loads.
    """
    if manufacturer == Manufacturer.ATI:
        return _parse_ati_pdf(file_bytes)
    elif manufacturer == Manufacturer.NEVADOS:
        return _parse_nevados_pdf(file_bytes)
    elif manufacturer == Manufacturer.NEXTPOWER:
        return _parse_nextpower_xlsx(file_bytes)
    return TOPLParseResult(
        success=False,
        manufacturer=manufacturer,
        errors=[f"Unknown manufacturer: {manufacturer}"],
        project_info=TOPLProjectInfo(),
    )


def topl_loads_to_session_dict(
    project_info: TOPLProjectInfo,
    loads: TOPLColumnLoads,
) -> dict[str, Any]:
    """Convert parsed TOPL data to session-state key-value pairs."""
    notes_parts = [f"Manufacturer: {project_info.manufacturer}"]
    if project_info.asce_version:
        notes_parts.append(f"ASCE: {project_info.asce_version}")
    if project_info.wind_speed_mph is not None:
        notes_parts.append(f"Wind: {project_info.wind_speed_mph} mph")
    if project_info.ground_snow_psf is not None:
        notes_parts.append(f"Snow: {project_info.ground_snow_psf} psf")
    if project_info.seismic_sds is not None:
        notes_parts.append(f"SDS: {project_info.seismic_sds}")
    notes_parts.append(f"Column type: {loads.column_type}")

    return {
        "project_name": project_info.project_name or "Untitled",
        "project_location": project_info.location,
        "project_notes": " | ".join(notes_parts),
        "dead_load": loads.dead_load,
        "live_load": 0.0,
        "snow_load": loads.snow_load,
        "wind_down": loads.wind_down,
        "wind_up": loads.wind_up,
        "wind_lateral": loads.wind_lateral,
        "wind_moment": loads.wind_moment,
        "seismic_lateral": loads.seismic_lateral,
        "seismic_vertical": loads.seismic_vertical,
        "seismic_moment": loads.seismic_moment,
        "lever_arm": loads.lever_arm,
    }


# ============================================================================
# Helpers
# ============================================================================

_NUM_RE = re.compile(r"[-\u2212]?\d+\.?\d*")


def _extract_floats(text: str) -> list[float]:
    """Pull all numeric values from *text*, treating Unicode minus as negative."""
    return [float(m.replace("\u2212", "-")) for m in _NUM_RE.findall(text)]


def _first_float(text: str, pattern: str, default: float | None = None) -> float | None:
    """Search *text* for *pattern*, return the first float found after the match."""
    m = re.search(pattern, text, re.IGNORECASE)
    if not m:
        return default
    rest = text[m.end():]
    nums = _extract_floats(rest[:80])
    return nums[0] if nums else default


def _first_str(text: str, pattern: str, default: str = "") -> str:
    m = re.search(pattern, text, re.IGNORECASE)
    return m.group(1).strip() if m else default


def _validate_loads(loads: TOPLColumnLoads, warnings: list[str]) -> None:
    """Add warnings for values outside typical solar pile ranges."""
    label = loads.column_type
    if loads.dead_load < 0:
        warnings.append(f"{label}: dead_load is negative ({loads.dead_load:.0f} lbs)")
    if loads.wind_lateral > 20000:
        warnings.append(f"{label}: wind_lateral={loads.wind_lateral:.0f} lbs seems high")
    if loads.lever_arm <= 0 or loads.lever_arm > 15:
        warnings.append(f"{label}: lever_arm={loads.lever_arm:.1f} ft outside typical 2-12 ft")


# ============================================================================
# ATI Parser
# ============================================================================

def _parse_ati_pdf(file_bytes: bytes) -> TOPLParseResult:
    """Parse ATI Ground Force Analysis PDF."""
    import pypdf

    warnings: list[str] = []
    errors: list[str] = []

    try:
        reader = pypdf.PdfReader(io.BytesIO(file_bytes))
    except Exception as e:
        return TOPLParseResult(
            success=False, manufacturer=Manufacturer.ATI,
            project_info=TOPLProjectInfo(), errors=[f"Failed to read PDF: {e}"],
        )

    pages = [p.extract_text() or "" for p in reader.pages]
    if len(pages) < 7:
        errors.append(f"Expected >= 7 pages in ATI TOPL, got {len(pages)}")
        return TOPLParseResult(
            success=False, manufacturer=Manufacturer.ATI,
            project_info=TOPLProjectInfo(), errors=errors,
        )

    # --- Project info from page 1 ---
    p1 = pages[0]
    info = TOPLProjectInfo(manufacturer="ATI")
    info.project_name = _first_str(p1, r"Project Name:\s*(.+?)(?:\n|$)")
    info.client = _first_str(p1, r"Client:\s*(.+?)(?:\n|$)")
    info.location = _first_str(p1, r"Location:\s*(.+?)(?:\n|$)")
    info.wind_speed_mph = _first_float(p1, r"Wind Speed \(mph\):")
    info.ground_snow_psf = _first_float(p1, r"Ground Snow \(psf\):")
    info.seismic_sds = _first_float(p1, r"Seismic Sds:")
    info.exposure = _first_str(p1, r"Terrain/Exposure:\s*(\S+)")
    info.risk_category = _first_str(p1, r"Building Category:\s*(.+?)(?:\n|$)")
    info.asce_version = _first_str(p1, r"Building Code:\s*(ASCE\s*[\d\-]+)")

    # Lever arm from max height of tracker
    lever_arm = _first_float(p1, r"Max height of tracker\s*=\s*", default=5.5)

    # --- Summary table (page 7 typically) ---
    summary_text = None
    pure_text = None
    for pt in pages:
        if "Max. Ground Force Summary Table" in pt:
            # Split into main table and pure interior table
            parts = pt.split("Pure Interior")
            summary_text = parts[0]
            if len(parts) > 1:
                pure_text = "Pure Interior" + parts[1]

    if not summary_text:
        errors.append("Could not find 'Max. Ground Force Summary Table'")
        return TOPLParseResult(
            success=False, manufacturer=Manufacturer.ATI,
            project_info=info, errors=errors,
        )

    # Parse heights from header
    heights = re.findall(r"HEIGHT\s*=\s*([\d.]+)", summary_text)
    h1 = float(heights[0]) if len(heights) > 0 else 5.5
    h2 = float(heights[1]) if len(heights) > 1 else 4.5

    # Parse each row of the main summary table
    # Format: label kips|k-ft <val1> <val2> <val3> <val4>
    row_patterns = [
        ("drag", r"\(drag\).*?kips"),
        ("wind_up", r"\(wind vertical up\).*?kips"),
        ("wind_down", r"\(wind vertical down\).*?kips"),
        ("snow", r"\(snow vertical down\).*?kips"),
        ("moment", r"\(ground-line\).*?k-ft"),
        ("dead", r"Dead-weight.*?kips"),
    ]

    table_data: dict[str, list[float]] = {}
    for key, pat in row_patterns:
        m = re.search(pat, summary_text, re.IGNORECASE)
        if m:
            rest = summary_text[m.end():]
            vals = _extract_floats(rest[:120])
            table_data[key] = vals[:4]  # Interior(H1), Exterior(H1), Interior(H2), Exterior(H2)
        else:
            warnings.append(f"ATI: Could not find row for '{key}'")
            table_data[key] = [0.0, 0.0, 0.0, 0.0]

    # Build column options: Interior/Exterior at each height
    column_configs = [
        (f"Interior (H={h1:.2f}')", 0),
        (f"Exterior (H={h1:.2f}')", 1),
        (f"Interior (H={h2:.2f}')", 2),
        (f"Exterior (H={h2:.2f}')", 3),
    ]

    loads_by_column: dict[str, TOPLColumnLoads] = {}
    for label, idx in column_configs:
        loads_by_column[label] = TOPLColumnLoads(
            column_type=label,
            wind_lateral=abs(table_data["drag"][idx]) * 1000,
            wind_up=abs(table_data["wind_up"][idx]) * 1000,
            wind_down=abs(table_data["wind_down"][idx]) * 1000,
            snow_load=abs(table_data["snow"][idx]) * 1000,
            wind_moment=abs(table_data["moment"][idx]) * 1000,
            dead_load=abs(table_data["dead"][idx]) * 1000,
            lever_arm=h1 if idx < 2 else h2,
        )

    # Parse Pure Interior table if present
    if pure_text:
        pure_data: dict[str, list[float]] = {}
        for key, pat in row_patterns:
            m = re.search(pat, pure_text, re.IGNORECASE)
            if m:
                rest = pure_text[m.end():]
                vals = _extract_floats(rest[:120])
                pure_data[key] = vals[:2]  # H1, H2
            else:
                pure_data[key] = [0.0, 0.0]

        for i, h in enumerate([h1, h2]):
            plabel = f"Pure Interior (H={h:.2f}')"
            if all(len(v) > i for v in pure_data.values()):
                loads_by_column[plabel] = TOPLColumnLoads(
                    column_type=plabel,
                    wind_lateral=abs(pure_data["drag"][i]) * 1000,
                    wind_up=abs(pure_data["wind_up"][i]) * 1000,
                    wind_down=abs(pure_data["wind_down"][i]) * 1000,
                    snow_load=abs(pure_data["snow"][i]) * 1000,
                    wind_moment=abs(pure_data["moment"][i]) * 1000,
                    dead_load=abs(pure_data["dead"][i]) * 1000,
                    lever_arm=h,
                )

    # --- Seismic from pages 8-9 ---
    seis_text = "\n".join(pages[7:])
    # N/S (weak-axis): V_per_column governs
    v_per_col = _first_float(seis_text, r"per column.*?=.*?=.*?([\d.]+)\s*kips", default=None)
    if v_per_col is None:
        # Try alternate pattern
        m = re.search(r"=\s*([\d.]+)\s*kips\s*\n.*Seismic shear", seis_text, re.IGNORECASE)
        if not m:
            m = re.search(r"\[\s*[\d.]+k\s*\]\s*=\s*([\d.]+)\s*kips", seis_text)
        if m:
            v_per_col = float(m.group(1))

    m_sc_match = re.search(r"𝑀𝑀𝑠𝑠𝑐𝑐\s*=.*?=\s*([\d.]+)\s*k-in", seis_text)
    m_sc = float(m_sc_match.group(1)) if m_sc_match else None

    if v_per_col is not None:
        seis_lat = v_per_col * 1000  # kips -> lbs
        seis_mom = (m_sc * 1000 / 12) if m_sc else (seis_lat * lever_arm)
        for loads in loads_by_column.values():
            loads.seismic_lateral = seis_lat
            loads.seismic_moment = seis_mom

    # Validate
    column_options = list(loads_by_column.keys())
    for loads in loads_by_column.values():
        _validate_loads(loads, warnings)

    return TOPLParseResult(
        success=True,
        manufacturer=Manufacturer.ATI,
        project_info=info,
        column_options=column_options,
        loads_by_column=loads_by_column,
        warnings=warnings,
        errors=errors,
    )


# ============================================================================
# Nevados Parser
# ============================================================================

def _parse_nevados_pdf(file_bytes: bytes) -> TOPLParseResult:
    """Parse Nevados Top of Pile Loads PDF."""
    import pypdf

    warnings: list[str] = []
    errors: list[str] = []

    try:
        reader = pypdf.PdfReader(io.BytesIO(file_bytes))
    except Exception as e:
        return TOPLParseResult(
            success=False, manufacturer=Manufacturer.NEVADOS,
            project_info=TOPLProjectInfo(), errors=[f"Failed to read PDF: {e}"],
        )

    pages = [p.extract_text() or "" for p in reader.pages]
    if not pages:
        return TOPLParseResult(
            success=False, manufacturer=Manufacturer.NEVADOS,
            project_info=TOPLProjectInfo(), errors=["PDF has no pages"],
        )

    # --- Page 1: Project info ---
    p1 = pages[0]
    info = TOPLProjectInfo(manufacturer="Nevados")
    info.project_name = _first_str(p1, r"Project Name:\s*(.+?)(?:\s*Date:|\n)")
    info.project_number = _first_str(p1, r"Project Number:\s*(.+?)(?:\s*Rev:|\n)")
    info.wind_speed_mph = _first_float(p1, r"Stow Design Wind Speed:")
    info.ground_snow_psf = _first_float(p1, r"Ground Snow Load:")
    info.seismic_sds = _first_float(p1, r"Seismic S.{0,3}:")
    info.exposure = _first_str(p1, r"Wind Exposure:\s*(\S+)")
    info.risk_category = _first_str(p1, r"Risk Category:\s*(\S+)")
    info.asce_version = _first_str(p1, r"ASCE Version:\s*(ASCE\s*[\d\-]+)")

    # Lever arm from Design Load Application Height (inches -> ft)
    # Non-Drive post height is typically the taller one
    app_height_in = _first_float(p1, r"Design Load Application Height\s*=\s*", default=None)
    if app_height_in is not None:
        lever_arm = app_height_in / 12.0
    else:
        lever_arm = 5.5
        warnings.append("Could not find Design Load Application Height; defaulting to 5.5 ft")

    # --- Pages 2+: Load tables ---
    # Combine all pages after page 1 for table parsing
    table_text = "\n".join(pages[1:])

    # Split into tables by "TABLE:" marker
    table_sections = re.split(r"TABLE:\s*", table_text)
    table_sections = [s for s in table_sections if s.strip()]

    if not table_sections:
        errors.append("No load tables found in Nevados TOPL")
        return TOPLParseResult(
            success=False, manufacturer=Manufacturer.NEVADOS,
            project_info=info, errors=errors,
        )

    loads_by_column: dict[str, TOPLColumnLoads] = {}
    column_options: list[str] = []
    _label_counter: dict[str, int] = {}  # Disambiguate duplicate table labels

    for section in table_sections:
        # Parse table metadata
        lines = section.strip().split("\n")
        table_num = _first_str(lines[0] if lines else "", r"^(\d+)")
        string_num = _first_str(section, r"String:\s*(\d+)")
        exposure = _first_str(section, r"Tracker Exposure:\s*(\w+)")
        load_case = _first_str(section, r"Load Case:\s*(.+?)(?:\n|$)")
        parts = [f"Table {table_num}"]
        if string_num:
            parts.append(f"Str {string_num}")
        if exposure:
            parts.append(exposure)
        if load_case:
            parts.append(load_case)
        base_label = " ".join(parts)
        # Disambiguate duplicate labels (e.g. multiple tracker configs)
        _label_counter[base_label] = _label_counter.get(base_label, 0) + 1
        if _label_counter[base_label] > 1:
            table_label = f"{base_label} ({_label_counter[base_label]})"
        else:
            table_label = base_label

        # Extract post info from bottom of table text
        # Posts appear as: "1 RE", "2 MS", "3 ND", "4 SD*", etc.
        post_pattern = re.compile(r"^\s*(\d+)\s+(RE|MS|ND|SD\*?|CTT)\s*$", re.MULTILINE)
        posts = post_pattern.findall(section)
        # Sort by post number — PDF text extraction may jumble column layout
        posts = sorted(set(posts), key=lambda x: int(x[0]))

        # Extract data rows: lines starting with Fh, Fv, or Ma followed by numbers
        data_row_re = re.compile(
            r"^(Fh|Fv|Ma)\s+([-\d.]+(?:\s+[-\d.]+)*)\s*(?:lbs|ft-lbs)?",
            re.MULTILINE,
        )
        data_rows = data_row_re.findall(section)

        if not data_rows:
            warnings.append(f"{table_label}: no data rows found")
            continue

        # Group into triplets (Fh, Fv, Ma) per post
        # Each post should have exactly 3 rows
        triplets = []
        current: dict[str, list[float]] = {}
        for load_type, nums_str in data_rows:
            vals = [float(x) for x in nums_str.split()]
            if len(vals) < 5:
                continue  # skip incomplete rows
            current[load_type] = vals
            if len(current) == 3:
                triplets.append(current)
                current = {}
        if current and len(current) == 3:
            triplets.append(current)

        # Assign post labels to triplets
        if not posts:
            # Generate generic labels
            posts = [(str(i + 1), "?") for i in range(len(triplets))]

        for i, triplet in enumerate(triplets):
            if i >= len(posts):
                break
            post_num, post_type = posts[i]
            label = f"{table_label} | Post {post_num} {post_type}"

            fh = triplet.get("Fh", [0] * 11)
            fv = triplet.get("Fv", [0] * 11)
            ma = triplet.get("Ma", [0] * 11)

            # Columns [0]=Wind-Uplift, [1]=Wind-Downforce, [2]=Dead, [3]=Snow, [4]=Ice
            wind_lat = max(abs(fh[0]), abs(fh[1])) if len(fh) >= 2 else 0
            wind_up_val = abs(fv[0]) if len(fv) >= 1 else 0  # uplift is negative Fv
            wind_down_val = abs(fv[1]) if len(fv) >= 2 else 0  # downforce is positive
            dead_val = abs(fv[2]) if len(fv) >= 3 else 0
            snow_val = abs(fv[3]) if len(fv) >= 4 else 0
            moment_val = max(abs(ma[0]), abs(ma[1])) if len(ma) >= 2 else 0

            loads = TOPLColumnLoads(
                column_type=label,
                dead_load=dead_val,
                snow_load=snow_val,
                wind_up=wind_up_val,
                wind_down=wind_down_val,
                wind_lateral=wind_lat,
                wind_moment=moment_val,
                lever_arm=lever_arm,
            )
            _validate_loads(loads, warnings)
            loads_by_column[label] = loads
            column_options.append(label)

    if not loads_by_column:
        errors.append("No load data extracted from any table")
        return TOPLParseResult(
            success=False, manufacturer=Manufacturer.NEVADOS,
            project_info=info, errors=errors,
        )

    return TOPLParseResult(
        success=True,
        manufacturer=Manufacturer.NEVADOS,
        project_info=info,
        column_options=column_options,
        loads_by_column=loads_by_column,
        warnings=warnings,
        errors=errors,
    )


# ============================================================================
# Nextpower Parser
# ============================================================================

def _parse_nextpower_xlsx(file_bytes: bytes) -> TOPLParseResult:
    """Parse Nextpower Top of Pile Loads Excel workbook."""
    import openpyxl

    warnings: list[str] = []
    errors: list[str] = []

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        return TOPLParseResult(
            success=False, manufacturer=Manufacturer.NEXTPOWER,
            project_info=TOPLProjectInfo(), errors=[f"Failed to read Excel: {e}"],
        )

    # --- REVISION sheet: project info ---
    info = TOPLProjectInfo(manufacturer="Nextpower")
    if "REVISION" in wb.sheetnames:
        rev = wb["REVISION"]
        info.project_name = str(rev["F38"].value or "")
        info.project_number = str(rev["F40"].value or "")
        info.client = str(rev["F41"].value or "")

    # --- Load sheets ---
    skip_sheets = {"Combined Excel", "REVISION"}
    load_sheets = [s for s in wb.sheetnames if s not in skip_sheets]

    if not load_sheets:
        errors.append("No load sheets found in Nextpower workbook")
        return TOPLParseResult(
            success=False, manufacturer=Manufacturer.NEXTPOWER,
            project_info=info, errors=errors,
        )

    loads_by_column: dict[str, TOPLColumnLoads] = {}
    column_options: list[str] = []

    for sheet_name in load_sheets:
        ws = wb[sheet_name]

        # Design parameters (from header rows)
        def _cell(addr: str) -> Any:
            return ws[addr].value

        asce_ver = _cell("O7")
        if asce_ver and not info.asce_version:
            info.asce_version = str(asce_ver)

        wind_speed = _safe_float(_cell("O9"))
        if wind_speed and info.wind_speed_mph is None:
            info.wind_speed_mph = wind_speed

        snow_psf = _safe_float(_cell("O10"))
        if snow_psf and info.ground_snow_psf is None:
            info.ground_snow_psf = snow_psf

        sds_val = _safe_float(_cell("AD16"))
        if sds_val and info.seismic_sds is None:
            info.seismic_sds = sds_val

        risk = _cell("O8")
        if risk and not info.risk_category:
            info.risk_category = str(risk)

        top_of_pier = _safe_float(_cell("V8")) or 6.0
        load_app_height = _safe_float(_cell("V9")) or top_of_pier

        # Seismic base shear total (lbs)
        seis_v_total = _safe_float(_cell("AD19")) or 0.0
        n_piers = _safe_float(_cell("V10")) or 7.0

        # --- Pier loads (rows 19-29) ---
        # Col M(13)=label, N(14)=Lateral, P(16)=Axial, R(18)=Moment(lbs-in),
        # U(21)=AbsDown, W(23)=AbsUp
        for row in range(19, 30):
            pier_label = ws.cell(row=row, column=13).value  # col M
            if pier_label is None:
                continue

            lateral = _safe_float(ws.cell(row=row, column=14).value)
            axial = _safe_float(ws.cell(row=row, column=16).value)
            moment_lbin = _safe_float(ws.cell(row=row, column=18).value)
            abs_down = _safe_float(ws.cell(row=row, column=21).value)
            abs_up = _safe_float(ws.cell(row=row, column=23).value)

            label = f"{sheet_name} | Pier {pier_label}"
            if label in loads_by_column:
                continue  # Skip duplicate pier rows (e.g. Motor pier)

            # Seismic per pier (evenly distributed)
            seis_per_pier = seis_v_total / max(n_piers, 1)

            loads = TOPLColumnLoads(
                column_type=label,
                wind_lateral=abs(lateral),
                wind_up=abs(abs_up),       # abs_up is negative in Nextpower
                wind_down=abs(abs_down),
                wind_moment=abs(moment_lbin) / 12.0,  # lbs-in -> ft-lbs
                lever_arm=load_app_height,
                seismic_lateral=seis_per_pier,
                seismic_moment=seis_per_pier * load_app_height,
            )

            # Try to extract dead and snow from detailed tables (rows 61+)
            _extract_nextpower_dead_snow(ws, pier_label, loads, warnings)

            _validate_loads(loads, warnings)
            loads_by_column[label] = loads
            column_options.append(label)

    if not loads_by_column:
        errors.append("No pier load data extracted from any sheet")
        return TOPLParseResult(
            success=False, manufacturer=Manufacturer.NEXTPOWER,
            project_info=info, errors=errors,
        )

    return TOPLParseResult(
        success=True,
        manufacturer=Manufacturer.NEXTPOWER,
        project_info=info,
        column_options=column_options,
        loads_by_column=loads_by_column,
        warnings=warnings,
        errors=errors,
    )


def _safe_float(val: Any) -> float:
    """Convert a cell value to float, returning 0.0 on failure."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        # Try extracting number from string like "105 mph"
        if isinstance(val, str):
            nums = _NUM_RE.findall(val)
            if nums:
                return float(nums[0].replace("\u2212", "-"))
        return 0.0


def _extract_nextpower_dead_snow(
    ws: Any,
    pier_label: Any,
    loads: TOPLColumnLoads,
    warnings: list[str],
) -> None:
    """Scan detailed load tables (rows 50+) for dead and snow per pier."""
    # The detailed section has columns: K=pier label, Q=DL, S=SL
    # Look for the pier in the detailed rows
    pier_str = str(pier_label)
    for row in range(50, min(ws.max_row + 1, 400)):
        cell_k = ws.cell(row=row, column=11).value  # col K
        if cell_k is not None and str(cell_k).strip() == pier_str:
            dl = _safe_float(ws.cell(row=row, column=17).value)  # col Q
            sl = _safe_float(ws.cell(row=row, column=19).value)  # col S
            if dl != 0.0 and loads.dead_load == 0.0:
                loads.dead_load = abs(dl)
            if sl != 0.0 and loads.snow_load == 0.0:
                loads.snow_load = abs(sl)
            if loads.dead_load > 0 and loads.snow_load > 0:
                break
